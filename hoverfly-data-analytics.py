# -*- coding: utf-8 -*-
"""
Created on Mon Sep  4 17:18:55 2017

Database functions for analysis of Andrew Lucus's hoverfly data

@author: CFord
"""

import csv
import sqlite3 as sql
import numpy
import simplekml
import matplotlib.pyplot as plt
from Bio import SeqIO
from openpyxl import load_workbook, Workbook
from collections import Counter
from itertools import combinations
from mpl_toolkits.basemap import Basemap

def bold_meta_index():
    """ (None -> {})
    
    Returns a list of indexes for the bold meta data for the following
    ids:
        boldid: <id>
        country: <country>
        lat: <latitude>
        lng: <longitude>
    """
    return {"boldid": 0, "country": 54, "lat": 46, "lng": 47 }

def andrew_bold_meta_index():
    """ (None -> {})
    
    Returns a list of indexes for the andrews bold meta data for the following
    ids:
        boldid: <id>
        country: <country>
        lat: <latitude>
        lng: <longitude>
    """
    return {"boldid": 0, "country": 4, "lat": 9, "lng": 10 }


def split_fasta_id(fastaid):
    """ (string) -> {}
    
    Takes a fasta id of the format <genus>_<species>_<boldid> and returns
    three variables as genus, species, boldid
    """
    parts = fastaid.split('_')
    if len(parts) > 3:
        raise ValueError('fasta id non-conformant "%s"' % fastaid)
    return parts[0],parts[1],parts[2]

def order_bold_ids(a,b):
    """ (string,string) -> string, string
    
    Given two bold ID's we return them ordered.
    """
    if a > b:
        a,b = b,a        
    return a,b

def create_db(db_name, recreate):
    """ (string, Bool) -> sqlite3.Connection
    
    If recreate is True then the current database will be deleted and the 
    tables will be recreated. If recreate is False then new tables will not
    be created.
    
    A connection to the database will be returned
    """
    con = sql.connect(db_name)
    if recreate:
        con.cursor().execute("DROP TABLE IF EXISTS sequences;")
        con.cursor().execute('''CREATE TABLE sequences(
                                boldid TEXT NOT NULL, 
                                genus TEXT NOT NULL, 
                                species TEXT NOT NULL, 
                                sequence TEXT,
                                country TEXT,
                                latitude REAL,
                                longitude REAL,
                                CONSTRAINT boldkey PRIMARY KEY (boldid));''')
        
        con.cursor().execute("DROP TABLE IF EXISTS distances;")
        con.cursor().execute('''CREATE TABLE distances(
                                boldid_a TEXT NOT NULL, 
                                boldid_b TEXT NOT NULL,
                                distance REAL NOT NULL,
                                CONSTRAINT bolddistkey PRIMARY KEY (boldid_a,boldid_b));''')
        
    return con

def import_fasta_into_db(fname,db):
    """ (string,sqlite3.Connection) -> None
    
    Assumes that the sequence ID's are formated as <genus>_<species>_<boldid> 
    and the file is a fasta formated file. Inserts data in to given db 
    connection
    """    
    cur = db.cursor()
    fd = open(fname,'r')
    for seq in SeqIO.parse(fd, "fasta"):
        genus,species,boldid = split_fasta_id(seq.id)
        cur.execute('INSERT INTO sequences VALUES(?,?,?,?,?,?,?);', (boldid,genus,species,str(seq.seq),"",0.0,0.0))
    db.commit()
    fd.close()

def import_country_origin_from_bold(fname,db,indexs):
    """ (string,sqlite3.Connection,{}) -> None
    
    Assumes that the file name points to a bold download that contains all
    the bold data. It will read in the data and for each bold id found it will
    update the database with the corrosponding country
    """
    cur = db.cursor()
    wb = load_workbook(filename = fname)
    sr = wb['bold-meta-data']
    bid_idx = indexs["boldid"]
    cnt_idx = indexs["country"]
    lat_idx = indexs["lat"]
    lng_idx = indexs["lng"]
    for ids in sr.iter_rows(row_offset=1):
        # {ID: species, lat, long, country}
        if ids[0].value != None:
            cur.execute('UPDATE sequences SET country = "%s", latitude = %f, longitude = %f WHERE boldid = "%s";' % (
                            ids[cnt_idx].value, 
                            float(ids[lat_idx].value if ids[lat_idx].value != None else 0.0 ), 
                            float(ids[lng_idx].value if ids[lng_idx].value != None else 0.0 ),
                            ids[bid_idx].value ))
    db.commit()

def import_distance_matrix_from_mega(fname,db):
    """ (string,sqlite3.Connection,{}) -> None
    
    Assumes file is in csv format, half matrix and was produced by mega.
    Inserts data in to distances table.
    """
    cur = db.cursor()
    with open(fname) as csvfile:
        boldids = []
        reader = csv.reader(csvfile)
        for row in reader:
            # A blank line will be the end
            if len(row) == 0:
                break
            genus,species,boldid = split_fasta_id(row[0])
            for index in range(0,len(boldids)):
                boldid_a,boldid_b = order_bold_ids(boldids[index],boldid)
                distance = abs(float(row[index+1]))
                if boldid_a > boldid_b:
                    boldid_a, boldid_b = boldid_b, boldid_a
                cur.execute('INSERT INTO distances VALUES(?,?,?);', (boldid_a,boldid_b,distance))
            boldids.append(boldid)
    db.commit()

def query_all(db):
    """ (sqlite3.Connection) -> []
    
    Returns a dictonary of boldids with sample info
    """
    cur = db.cursor()
    cur.execute('select * from sequences;')
    all = cur.fetchall()
    return all

def query_all_lat_lngs(db):
    """ (sqlite3.Connection) -> [][]
    
    Returns a list of lat and a list of lngs
    """
    cur = db.cursor()
    cur.execute('select latitude,longitude from sequences;')
    all = cur.fetchall()
    ret = [i for i in all if i[0] != 0 and i[1] != 0]
    return zip(*ret)

def query_unique_species(db):
    """ (sqlite3.Connection) -> []
    
    Returns a list of unique species.
    """
    cur = db.cursor()
    cur.execute('select genus,species from sequences group by genus,species;')
    all_species = cur.fetchall()
    unique = set()
    for spp in all_species:
        unique.add(' '.join(spp))
    return unique

def query_unique_genus(db):
    """ (sqlite3.Connection) -> []
    
    Returns a list of unique genus.
    """
    cur = db.cursor()
    cur.execute('select genus from sequences group by genus;')
    all_species = cur.fetchall()
    unique = set()
    for spp in all_species:
        unique.add(' '.join(spp))
    return unique

def genus_unique_species(genus,db):
    """ (string, sqlite3.Connection) -> []
    
    Given a genus, return a list of unique species within that genus
    """
    cur = db.cursor()
    cur.execute('select genus,species from sequences where genus = "%s" group by species;' 
                % (genus))
    all_species = cur.fetchall()
    unique = set()
    for spp in all_species:
        unique.add(' '.join(spp))
    return unique

def query_genus_with_more_than_one_species(db):
    """ (sqlite3.Connection) -> []
    
    Returns a list of genus with more than one species.
    """
    cur = db.cursor()
    cur.execute('select genus,species from sequences group by genus,species;')
    all_species = cur.fetchall()
    genus_count = Counter(gen[0] for gen in all_species)
    return [k for k, v in genus_count.items() if v > 1]

def query_species_boldids(spp,db):
    """ (string, sqlite3.Connection) -> []
    
    Given a spp as i.e. 'Syrphus rectus' and a the database connection
    return a list of boldids where the boldid matches that species
    """
    genus,species = spp.split(' ')
    cur = db.cursor()
    cur.execute('select boldid from sequences where genus = "%s" and species = "%s";' % (genus,species))
    return [i[0] for i in cur.fetchall()] 

def query_genus_boldids(genus,db):
    """ (string, sqlite3.Connection) -> []
    
    Given a genus as i.e. 'Syrphus' and a the database connection
    return a list of boldids where the boldid matches that genus
    """
    cur = db.cursor()
    cur.execute('select boldid from sequences where genus = "%s";' % (genus))
    return [i[0] for i in cur.fetchall()] 

def query_not_species_boldids(spp,db):
    """ (string, sqlite3.Connection) -> []
    
    Given a spp as i.e. 'Syrphus rectus' and a the database connection
    return a list of boldids where the boldid doesn't match that species
    """
    genus,species = spp.split(' ')
    cur = db.cursor()
    cur.execute('select boldid from sequences where not (genus = "%s" and species = "%s");' % (genus,species))
    return [i[0] for i in cur.fetchall()] 

def query_subset_species_boldids(spp,db,like='',not_like=False):
    """ (string, sqlite3.Connection,string) -> []
    
    Given a spp as i.e. 'Syrphus rectus' and a the database connection
    return a list of boldids where the boldid matches that species with any
    like clause if it is given applied to the boldid.
    """
    genus,species = spp.split(' ')
    cur = db.cursor()
    if like == '':
        cur.execute('select boldid from sequences where genus = "%s" and species = "%s";' % (genus,species))
    else:
        if not_like:
            cur.execute('select boldid from sequences where genus = "%s" and species = "%s" and boldid not like "%s";' 
                        % (genus,species,like))
        else:
            cur.execute('select boldid from sequences where genus = "%s" and species = "%s" and boldid like "%s";' 
                        % (genus,species,like))
    return [i[0] for i in cur.fetchall()] 

def query_min_max_distance(db):
    """ (sqlite3.Connection) -> float,float
    
    Returns the min and max distances within the distances table.
    """
    cur = db.cursor()
    cur.execute('select min(distance),max(distance) from distances;')
    data = cur.fetchall()
    return data[0][0],data[0][1]

def query_distances_for_boldids(bids,db):
    """ ([],sqlite3.Connection) -> []
    
    Given a list of bold ids return a list of distances for bold ids 
    matching those given.
    """
    boldids_string = ','.join('"{0}"'.format(w) for w in bids)
    cur = db.cursor()
    cur.execute('select distance from distances where (boldid_a in (%s) or boldid_b in (%s));' % (boldids_string, boldids_string))
    return [i[0] for i in cur.fetchall()] 

def query_distances_for_not_boldids(bids,db):
    """ ([],sqlite3.Connection) -> []
    
    Given a list of bold ids return a list of distances for bold ids not 
    matching those given.
    """
    boldids_string = ','.join('"{0}"'.format(w) for w in bids)
    cur = db.cursor()
    cur.execute('select distance from distances where (boldid_a not in (%s) and boldid_b not in (%s));' % (boldids_string, boldids_string))
    return [i[0] for i in cur.fetchall()] 

def query_distances_for_boldids_one_sided(bids,db):
    """ ([],sqlite3.Connection) -> []
    
    Given a list of bold ids return a list of distances for bold ids where
    the boldids match only one of the pair in the distance table, discard any
    matching both.
    """
    boldids_string = ','.join('"{0}"'.format(w) for w in bids)
    cur = db.cursor()
    cur.execute("""select distance from distances where 
                   (boldid_a in (%s) and boldid_b not in (%s)) 
                   or
                   (boldid_a not in (%s) and boldid_b in (%s));""" % (boldids_string,boldids_string,boldids_string,boldids_string))
    return [i[0] for i in cur.fetchall()] 

def query_distances_for_boldids_on_both_sides(bids,db):
    """ ([],sqlite3.Connection) -> []
    
    Given a list of bold ids return a list of distances for bold ids where
    the boldids match both sides of the distance pair, discard any where
    only one side matches a bold id given.
    """
    boldids_string = ','.join('"{0}"'.format(w) for w in bids)
    cur = db.cursor()
    cur.execute("""select distance from distances where 
                   (boldid_a in (%s) and boldid_b in (%s));""" % (boldids_string,boldids_string))
    return [i[0] for i in cur.fetchall()] 

def query_distances_for_boldids_on_given_sides(bids1,bids2,db):
    """ ([],[],sqlite3.Connection) -> []
    
    Given two lists of boldids 1,2 give back a list where bold ids match on both
    sides but don't match within both sides so if we have
    
    bids1 = 1A, 1B
    bids2 = 2A, 2B
    
    We should only give back data where:
    1A - 2A
    1A - 2B
    1B - 2A
    1B - 2B
    
    or swapped around.
    """
    boldids_string1 = ','.join('"{0}"'.format(w) for w in bids1)
    boldids_string2 = ','.join('"{0}"'.format(w) for w in bids2)
    cur = db.cursor()
    cur.execute("""select distance from distances where 
                   (boldid_a in (%s) and boldid_b in (%s)) or (boldid_a in (%s) and boldid_b in (%s));""" 
                   % (boldids_string1,boldids_string2,boldids_string2,boldids_string1))
    return [i[0] for i in cur.fetchall()] 

def query_distance_for_spp_combination(spp_a,spp_b,db):
    """ (string,string,sqlite3.Connection) -> []
    
    Given a two species return the list of distances for species a against
    species b but not a <-> a or b <-> b.
    """
    bid_a_string = ",".join('"{0}"'.format(w) for w in query_species_boldids(spp_a,db))
    bid_b_string = ",".join('"{0}"'.format(w) for w in query_species_boldids(spp_b,db))
    cur = db.cursor()
    cur.execute("""select distance from distances where 
                   (boldid_a in (%s) and boldid_b in (%s)) 
                   or 
                   (boldid_a in (%s) and boldid_b in (%s));""" 
                   % (bid_a_string,bid_b_string,bid_b_string,bid_a_string))
    return [i[0] for i in cur.fetchall()]

def query_bins(number_of_bins,db):
    """ (int,sqlite3.Connection) -> numpy.array
    
    Returns the bins as a list for binning the distances
    """
    minb,maxb = query_min_max_distance(db)
    bin_size = (maxb - minb) / number_of_bins
    return numpy.arange(minb,maxb,bin_size)

def histogram_inter_species(spp,num_bins,db,distance_as_percentage=False,subset_like='',not_subset=False):
    """ (string,int,sqlite3.Connection,string) -> frequency,bins
    
    Returns a histogram of frequency,bins for the interspecic counts that fall
    within the num_bins calculated of species given to all the other species.
    If subset_like is not empty then select on the boldids and for species
    that meet the boldid like pattern.
    """
    bold_ids = query_subset_species_boldids(spp,db,like=subset_like,not_like=not_subset)
    distances = numpy.array(query_distances_for_boldids_one_sided(bold_ids,db))
    if distance_as_percentage:
        distances = numpy.ceil(distances*100)
        disbins = numpy.around(query_bins(num_bins,db)*100,0)
    else:
        disbins = query_bins(num_bins,db)
    return numpy.histogram(distances,disbins)

def histogram_intra_species(spp,num_bins,db,distance_as_percentage=False,subset_like='',not_subset=False):
    """ (string,int,sqlite3.Connection) -> frequency,bins
    
    Returns a histogram of frequency,bins for the intraspecic counts that fall
    within the num_bins calculated of species given to all species of that
    type. If subset_like is not empty then select on the boldids and for 
    species that meet the boldid like pattern.
    """
    bold_ids = query_subset_species_boldids(spp,db,like=subset_like,not_like=not_subset)
    distances = numpy.array(query_distances_for_boldids_on_both_sides(bold_ids,db))
    if distance_as_percentage:
        distances = numpy.ceil(distances*100)
        disbins = numpy.around(query_bins(num_bins,db)*100,0)
    else:
        disbins = query_bins(num_bins,db)
    return numpy.histogram(distances,disbins)

def histogram_inter_all_species(num_bins,db,distance_as_percentage=False,subset_like='',not_subset=False):
    """ (int,sqlite3.Connection) -> frequency,bins
    
    Calculates the histogram for all species for the interspecific distance. If
    a like clause is given then only the boldids matching that like clause and
    the species will be selected for.
    """
    freq_hist = numpy.zeros(num_bins-1)
    for spp in query_unique_species(db):
        freq,bins = histogram_inter_species(spp,num_bins,db,distance_as_percentage,subset_like,not_subset)
        freq_hist += freq 
    return freq_hist,bins

def histogram_intra_all_species(num_bins,db,distance_as_percentage=False,subset_like='',not_subset=False):
    """ (int,sqlite3.Connection) -> frequency,bins
    
    Calculates the histogram for all species for the intraspecific distance. If
    a like clause is given then only the boldids matching that like clause and
    the species will be selected for.
    """
    freq_hist = numpy.zeros(num_bins-1)
    for spp in query_unique_species(db):
        freq,bins = histogram_intra_species(spp,num_bins,db,distance_as_percentage,subset_like,not_subset)
        freq_hist += freq 
    return freq_hist,bins

def histogram_as_percentage(hist):
    """ (numpy.array)
    
    Returns the given histogram as a percentage frequency instead of raw frequency
    """
    return numpy.around(numpy.true_divide(hist, hist.sum(keepdims=True)) * 100, 0)

def mean_std_intra_each_species(db,subset_like='',not_subset=False):
    """ (int,sqlite3.Connection) -> {}
    
    Calculates the mean and std for each spp distance set
    """
    data = {}
    for spp in query_unique_species(db):
        bold_ids = query_subset_species_boldids(spp,db,like=subset_like,not_like=not_subset)
        distances = query_distances_for_boldids_on_both_sides(bold_ids,db)
        data[spp] = {"mean": numpy.mean(distances) if len(distances) > 0 else 0, 
                     "std": numpy.std(distances) if len(distances) > 0 else 0,
                     "count": len(bold_ids)}
        
    return data

def mean_std_intra_sual_vs_row(db):
    """ (int,sqlite3.Connection) -> {}
    
    Calculates the mean and std for each spp with more than 1 speciemen
    in SAUL against the same spp in the RoW with more than 1 speciemen
    
    Species X with (A,B) samples in SUAL
    Species X with (1,2) samples in RoW
    
    A - 1 distance
    A - 2 distance
    B - 1 distance
    B - 2 distance
    
    return mean, std of distance
    """
    data = {}
    for spp in query_unique_species(db):
        sual_bold_ids = query_subset_species_boldids(spp,db,like="SUAL%",not_like=False)
        row_bold_ids = query_subset_species_boldids(spp,db,like="SUAL%",not_like=True)
        if len(sual_bold_ids) <= 1 or len(row_bold_ids) <= 1:
            data[spp] = {"mean": 0, 
                         "std": 0,
                         "count SUAL": len(sual_bold_ids),
                         "count RoW": len(row_bold_ids)}
            continue
        else:
            print(spp,len(sual_bold_ids),len(row_bold_ids))
            distances = query_distances_for_boldids_on_given_sides(sual_bold_ids,row_bold_ids,db)
            data[spp] = {"mean": numpy.mean(distances) if len(distances) > 0 else 0, 
                         "std": numpy.std(distances) if len(distances) > 0 else 0,
                         "count SUAL": len(sual_bold_ids),
                         "count RoW": len(row_bold_ids)}
    return data

def mean_std_intra_each_genus_with_more_than_one_species(db):
    """ (int,sqlite3.Connection) -> {}
    
    Calculates the mean and std within each genus with more than one species
    """
    data = {}
    for genus in query_genus_with_more_than_one_species(db):
        species = genus_unique_species(genus,db)
        distances = []
        for spp_pair in combinations(species,2):
            distances.extend(query_distance_for_spp_combination(spp_pair[0],spp_pair[1],db))
        data[genus] = {"mean": numpy.mean(distances) if len(distances) > 0 else 0, 
                       "std": numpy.std(distances) if len(distances) > 0 else 0,
                       "count": len(species)}
        
    return data

def mean_difference_between_all_species(db):
    """ (sqllite3.Connection) --> {}
    
    Calculates the mean differences between all species.
    Take all individuals of species A and species B and calculate the mean.
    """
    data = {}
    for spp_pair in combinations(query_unique_species(db),2):
        data[spp_pair] = numpy.mean(query_distance_for_spp_combination(spp_pair[0],spp_pair[1],db))
                
    return data

def mean_difference_between_all_species_between(lower,upper,db):
    """ (string,string,float,float,sqlite3.Connection) -> []
    
    Calculates the mean differences between all species.
    Take all individuals of species A and species B and calculate the mean.
    Filter the distance to thoes only > lower and <= upper
    """
    all_dists = mean_difference_between_all_species(db)
    return {k: v for k,v in all_dists.items() if v > 0.00 and v <= 0.03}

def output_txt_location_file(fname,db):
    """ (string,sqlite3.Connection) -> None
    
    Creates a txt file with location information in it
    """
    fd = open(fname,'w')
    fd.write('Name,Latitude,Longitude\n')
    all = query_all(db)
    for row in all:
        bid = row[0]
        lat = row[5]
        lng = row[6]
        if lat == 0 or lng == 0:
            continue
        fd.write('%s,%f,%f\n' % (bid,lat,lng))
    fd.close()

def output_kml_file(fname,db):
    """ (string,sqlite3.Connection) -> None
    
    Creates a KML file from the location data
    """
    kml = simplekml.Kml()
    all = query_all(db)
    folders = {}
    for row in all:
        spp = row[1] + " " + row[2]
        fol = None
        if not spp in folders:
            fol = kml.newfolder(name=spp)
            folders[spp] = fol
        else:
            fol = folders[spp]
        
        bid = row[0]
        cnt = row[4]
        lat = row[5]
        lng = row[6]
        
        if lat == None or lng == None or cnt == None or cnt == '':
            print('No geo:',bid,spp,cnt)
        else:
            fol.newpoint(name='%s,%s' % (bid,cnt) , coords=[(lng,lat)])
    kml.save(fname)
    
def output_stats_excel_file(fname):
    """ (string) -> None
    
    Creates a file with all of the stats in
    """
    wb = Workbook()
    
    ws = wb.active
    ws.title = 'All Inter vs Intra'
    ws.cell(row=1, column=1, value="distance %")
    ws.cell(row=1, column=2, value="inter")
    ws.cell(row=1, column=3, value="intra")
    ws.cell(row=1, column=4, value="inter %")
    ws.cell(row=1, column=5, value="intra %")
    
    ws2 = wb.create_sheet('All Spp Intra stats')
    ws2.cell(row=1, column=1, value="species")
    ws2.cell(row=1, column=2, value="mean")
    ws2.cell(row=1, column=3, value="std")
    ws2.cell(row=1, column=4, value="count")
    
    ws3 = wb.create_sheet('Only SUAL, Inter vs Intra')
    ws3.cell(row=1, column=1, value="distance %")
    ws3.cell(row=1, column=2, value="inter")
    ws3.cell(row=1, column=3, value="intra")
    ws3.cell(row=1, column=4, value="inter %")
    ws3.cell(row=1, column=5, value="intra %")
    
    ws4 = wb.create_sheet('Only SUAL, Spp Intra stats')
    ws4.cell(row=1, column=1, value="species")
    ws4.cell(row=1, column=2, value="mean")
    ws4.cell(row=1, column=3, value="std")
    ws4.cell(row=1, column=4, value="count")

    ws5 = wb.create_sheet('Without SUAL, Inter vs Intra')
    ws5.cell(row=1, column=1, value="distance %")
    ws5.cell(row=1, column=2, value="inter")
    ws5.cell(row=1, column=3, value="intra")
    ws5.cell(row=1, column=4, value="inter %")
    ws5.cell(row=1, column=5, value="intra %")
    
    ws6 = wb.create_sheet('Without SUAL, Spp Intra stats')
    ws6.cell(row=1, column=1, value="species")
    ws6.cell(row=1, column=2, value="mean")
    ws6.cell(row=1, column=3, value="std")
    ws6.cell(row=1, column=4, value="count")
    
    ws7 = wb.create_sheet('Genus with more than on spp')
    ws7.cell(row=1, column=1, value="genus")
    ws7.cell(row=1, column=2, value="mean")
    ws7.cell(row=1, column=3, value="std")
    ws7.cell(row=1, column=4, value="count")
    
    ws8 = wb.create_sheet('Spp pairs where dist...')
    ws8.cell(row=1, column=1, value="species A")
    ws8.cell(row=1, column=2, value="species B")
    ws8.cell(row=1, column=3, value="mean distance > 0.00 && <= 0.03")
    
    ws9 = wb.create_sheet('Intra spp SUAL vs RoW')
    ws9.cell(row=1, column=1, value="spp")
    ws9.cell(row=1, column=2, value="mean")
    ws9.cell(row=1, column=3, value="std")
    ws9.cell(row=1, column=4, value="count SUAL")
    ws9.cell(row=1, column=5, value="count RoW")
    
    db = open_db()
    
    # Global
    hist_intra,bins_intra = histogram_intra_all_species(27,db,True)
    hist_inter,bins_inter = histogram_inter_all_species(27,db,True)
    hist_intra_percent    = histogram_as_percentage(hist_intra)
    hist_inter_percent    = histogram_as_percentage(hist_inter)

    mean_std_overall = mean_std_intra_each_species(db)
    
    row_index = 2
    for value in zip(bins_intra,hist_inter,hist_intra,hist_inter_percent,hist_intra_percent):
        ws.cell(row=row_index, column=1, value=value[0])
        ws.cell(row=row_index, column=2, value=value[1])
        ws.cell(row=row_index, column=3, value=value[2])
        ws.cell(row=row_index, column=4, value=value[3])
        ws.cell(row=row_index, column=5, value=value[4])
        row_index += 1
        
    row_index = 2
    for spp in mean_std_overall:
        ws2.cell(row=row_index, column=1, value="%s" % spp)
        ws2.cell(row=row_index, column=2, value=mean_std_overall[spp]["mean"])
        ws2.cell(row=row_index, column=3, value=mean_std_overall[spp]["std"])
        ws2.cell(row=row_index, column=4, value=mean_std_overall[spp]["count"])
        row_index += 1

    # Only SUAL
    hist_intra,bins_intra = histogram_intra_all_species(27,db,True,'SUAL%',False)
    hist_inter,bins_inter = histogram_inter_all_species(27,db,True,'SUAL%',False)
    hist_intra_percent    = histogram_as_percentage(hist_intra)
    hist_inter_percent    = histogram_as_percentage(hist_inter)

    mean_std_overall = mean_std_intra_each_species(db,'SUAL%',False)
    
    row_index = 2
    for value in zip(bins_intra,hist_inter,hist_intra,hist_inter_percent,hist_intra_percent):
        ws3.cell(row=row_index, column=1, value=value[0])
        ws3.cell(row=row_index, column=2, value=value[1])
        ws3.cell(row=row_index, column=3, value=value[2])
        ws3.cell(row=row_index, column=4, value=value[3])
        ws3.cell(row=row_index, column=5, value=value[4])
        row_index += 1
        
    row_index = 2
    for spp in mean_std_overall:
        ws4.cell(row=row_index, column=1, value="%s" % spp)
        ws4.cell(row=row_index, column=2, value=mean_std_overall[spp]["mean"])
        ws4.cell(row=row_index, column=3, value=mean_std_overall[spp]["std"])
        ws4.cell(row=row_index, column=4, value=mean_std_overall[spp]["count"])
        row_index += 1
        
    # Without SUAL
    hist_intra,bins_intra = histogram_intra_all_species(27,db,True,'SUAL%',True)
    hist_inter,bins_inter = histogram_inter_all_species(27,db,True,'SUAL%',True)
    hist_intra_percent    = histogram_as_percentage(hist_intra)
    hist_inter_percent    = histogram_as_percentage(hist_inter)

    mean_std_overall = mean_std_intra_each_species(db,'SUAL%',True)
    
    row_index = 2
    for value in zip(bins_intra,hist_inter,hist_intra,hist_inter_percent,hist_intra_percent):
        ws5.cell(row=row_index, column=1, value=value[0])
        ws5.cell(row=row_index, column=2, value=value[1])
        ws5.cell(row=row_index, column=3, value=value[2])
        ws5.cell(row=row_index, column=4, value=value[3])
        ws5.cell(row=row_index, column=5, value=value[4])
        row_index += 1
        
    row_index = 2
    for spp in mean_std_overall:
        ws6.cell(row=row_index, column=1, value="%s" % spp)
        ws6.cell(row=row_index, column=2, value=mean_std_overall[spp]["mean"])
        ws6.cell(row=row_index, column=3, value=mean_std_overall[spp]["std"])
        ws6.cell(row=row_index, column=4, value=mean_std_overall[spp]["count"])
        row_index += 1
        
    # Genus with more than 
    genus_wmtos = mean_std_intra_each_genus_with_more_than_one_species(db)
    
    row_index = 2
    for genus in genus_wmtos:
        ws7.cell(row=row_index, column=1, value="%s" % genus)
        ws7.cell(row=row_index, column=2, value=genus_wmtos[genus]["mean"])
        ws7.cell(row=row_index, column=3, value=genus_wmtos[genus]["std"])
        ws7.cell(row=row_index, column=4, value=genus_wmtos[genus]["count"])
        row_index += 1
        
    # Species pairs where all of A distance to B > 0.00 && <= 0.03 
    spp_diff = mean_difference_between_all_species_between(0.00,0.03,db)
    
    row_index = 2
    for key,val in spp_diff.items():
        ws8.cell(row=row_index, column=1, value="%s" % key[0])
        ws8.cell(row=row_index, column=2, value="%s" % key[1])
        ws8.cell(row=row_index, column=3, value=val)
        row_index += 1
    
    # By species SUAL vs Row
    sual_vs_row = mean_std_intra_sual_vs_row(db)

    row_index = 2
    for spp in sual_vs_row:
        ws9.cell(row=row_index, column=1, value="%s" % spp)
        ws9.cell(row=row_index, column=2, value=sual_vs_row[spp]["mean"])
        ws9.cell(row=row_index, column=3, value=sual_vs_row[spp]["std"])
        ws9.cell(row=row_index, column=4, value=sual_vs_row[spp]["count SUAL"])
        ws9.cell(row=row_index, column=5, value=sual_vs_row[spp]["count RoW"])
        row_index += 1
    
    wb.save(fname)
    db.close()

def query_test(db):
    """ (sqlite3.Connection) -> None
    
    Runs a query to test the database
    """
    cur = db.cursor()
    cur.execute('select * from sequences where boldid = "NORSY287-12";')
    print(cur.fetchall())
    cur.execute('select * from sequences where boldid = "SUAL116-16";')
    print(cur.fetchall())
    cur.execute('select * from distances where boldid_a = "NORSY174-12" and boldid_b = "NORSY371-12";')
    print(cur.fetchall())

def create_and_test_db():
    """ (None) -> sqlite3.Connection
    
    Recreates the database and tests two selects. Returns the open database
    """
    db = create_db("andrews-sequence-database.db",True)
    import_fasta_into_db('andrew-fasta-data-alignment.fas',db)
    import_country_origin_from_bold('bold-meta-data.xlsx',db,bold_meta_index())
    import_country_origin_from_bold('andrew_bold_data.xlsx',db,andrew_bold_meta_index())
    import_distance_matrix_from_mega("andrew-seq-distance-matrix.csv",db)
    query_test(db)
    return db

def open_db():
    """ (None) -> sqlite3.Connection
    
    Returns the open database
    """
    return create_db("andrews-sequence-database.db",False)

def draw_map_of_samples(db):
    """ (None) -> sqlite3.Connection
    
    Using the lat/lngs of each sample draw a map with them pinpointed
    """
    lats,lons = query_all_lat_lngs(db)
    m = Basemap(projection='merc',llcrnrlat=-60,urcrnrlat=80,\
            llcrnrlon=-180,urcrnrlon=180,lat_ts=40,resolution='c')
    
    #m.drawcoastlines()
    #m.shadedrelief()
    m.drawcoastlines()
    m.drawcountries()
    m.fillcontinents(color = 'white')
    m.drawmapboundary()
    
    x,y = m(lons, lats)
    m.plot(x, y, 'bo', markersize=3)
    
    plt.savefig("andrew_samples.pdf", dpi=600, format='pdf')